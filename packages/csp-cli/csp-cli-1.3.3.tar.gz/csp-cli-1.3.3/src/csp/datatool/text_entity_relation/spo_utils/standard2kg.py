# -*- coding: utf-8 -*-
"""
Created on 2022-4-13
@author: cxy
"""

import json
import os
import sys

import openpyxl
import pandas as pd
# import argparse

xls_path = None
csv_path = None


def get_st_table(id2cid, id2value, cid2cname):
    col_list = ['uid', 'label', 'object_value']

    # 生成实体类型中文名到英文名的映射(字典cname2ename)
    cname2ename = {}
    for k, v in cid2cname.items():
        new_key = v
        new_value = 'label' + str(k)
        cname2ename[new_key] = new_value

    # 生成'实体'表格
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('实体数据')
    ws_default = wb['Sheet']
    wb.remove(ws_default)
    for index, item in enumerate(col_list):
        ws.cell(row=1, column=index + 1).value = item
    row = 2
    for k, v in id2value.items():
        ws.cell(row=row, column=1).value = str(k)  # int to str 20220623
        ws.cell(row=row, column=3).value = v
        new_key = cid2cname[str(id2cid[k])]
        # new_key = cid2cname[str(id2cid[k])]
        ws.cell(row=row, column=2).value = cname2ename[new_key]
        row += 1

    # 输出保存
    # wb.save('./xlsx/实体数据.xlsx')
    # data_xls = pd.read_excel('./xlsx/实体数据.xlsx', index_col=0)
    # data_xls.to_csv('./csv/实体数据.csv', encoding='utf-8')

    save_path_xlsx = os.path.join(xls_path, '实体数据.xlsx')
    wb.save(save_path_xlsx)
    data_xls = pd.read_excel(save_path_xlsx, index_col=0)
    save_path_xlsx = os.path.join(csv_path, '实体数据.csv')
    data_xls.to_csv(save_path_xlsx, encoding='utf-8')
     
    print('实体数据.csv：',save_path_xlsx)
    return cname2ename


def get_gx_table(id2all):
    col_list = ['uid', 'label', 'from', 'to', 'fromLabel', 'toLabel']

    # 生成'关系'表格
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('关系数据')
    ws_default = wb['Sheet']
    wb.remove(ws_default)
    for index, item in enumerate(col_list):
        ws.cell(row=1, column=index + 1).value = item
    row = 2
    for k, v in id2all.items():
        ws.cell(row=row, column=1).value = str(k)  # int to str 20220623
        ws.cell(row=row, column=2).value = v[4]
        ws.cell(row=row, column=3).value = v[0]
        ws.cell(row=row, column=4).value = v[1]
        ws.cell(row=row, column=5).value = v[2]
        ws.cell(row=row, column=6).value = v[3]
        row += 1

    # 输出保存
    # wb.save('./xlsx/关系数据.xlsx')
    # data_xls = pd.read_excel('./xlsx/关系数据.xlsx', index_col=0)
    # data_xls.to_csv('./csv/关系数据.csv', encoding='utf-8')

    save_path_xlsx = os.path.join(xls_path, '关系数据.xlsx')
    wb.save(save_path_xlsx)
    data_xls = pd.read_excel(save_path_xlsx, index_col=0)
    save_path_xlsx = os.path.join(csv_path, '关系数据.csv')
    data_xls.to_csv(save_path_xlsx, encoding='utf-8')

    print('关系数据.csv：',save_path_xlsx)

def get_bt_table(cname2ename, connect_list):
    col_list = ['objectType', 'ename', 'cname', 'from_node', 'to_node']

    # 生成'本体'表格
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('本体设计')
    ws_default = wb['Sheet']
    wb.remove(ws_default)
    for index, item in enumerate(col_list):
        ws.cell(row=1, column=index + 1).value = item

    row = 2  # 实体部分开始
    for k, v in cname2ename.items():
        ws.cell(row=row, column=1).value = 0
        ws.cell(row=row, column=2).value = v
        ws.cell(row=row, column=3).value = k
        row += 1

    row = 2 + len(cname2ename)  # 关系部分开始
    for item in connect_list:
        ws.cell(row=row, column=1).value = 1
        ws.cell(row=row, column=2).value = item[0]
        ws.cell(row=row, column=3).value = item[1]
        ws.cell(row=row, column=4).value = item[2]
        ws.cell(row=row, column=5).value = item[3]
        row += 1

    # 输出保存
    # wb.save('./xlsx/本体设计.xlsx')
    # data_xls = pd.read_excel('./xlsx/本体设计.xlsx', index_col=0)
    # data_xls.to_csv('./csv/本体设计.csv', encoding='utf-8')

    save_path_xlsx = os.path.join(xls_path, '本体设计.xlsx')
    wb.save(save_path_xlsx)
    data_xls = pd.read_excel(save_path_xlsx, index_col=0)
    save_path_xlsx = os.path.join(csv_path, '本体设计.csv')
    data_xls.to_csv(save_path_xlsx, encoding='utf-8')
    print('本体设计.csv：',save_path_xlsx)


def handle_json_file(label_path, connect_path, label_path_c, connect_path_c, output):
    global xls_path, csv_path
    xls_path = os.path.join(output, "xlsx") if output else './xlsx'
    csv_path = os.path.join(output, "csv") if output else './csv'
    os.makedirs(xls_path, exist_ok=True)
    os.makedirs(csv_path, exist_ok=True)
    # 创建表格保存目录
    # xls_path = './xlsx'
    # if not os.path.exists(xls_path):
    #     os.mkdir(xls_path)
    # csv_path = './csv'
    # if not os.path.exists(csv_path):
    #     os.mkdir(csv_path)

    # 处理labels.json文件
    label_id2cid = {}
    label_id2value = {}
    with open(label_path, "r", encoding="utf-8") as f:
        json_list = json.load(f)
    for item_dict in json_list:
        key = str(item_dict['srcId']) + '-' + str(item_dict['id'])  # modify by 20220627
        label_id2cid[key] = str(item_dict['categoryId'])  # modify by 20220627
        label_id2value[str(key)] = item_dict['name']

    # 处理labelCategories.json文件
    label_cid2cname = {}
    with open(label_path_c, "r", encoding="utf-8") as f:
        json_list = json.load(f)
    for item_dict in json_list:
        key = item_dict['id']
        label_cid2cname[str(key)] = item_dict['text']

    # 生成‘实体’表格, 并返回字典:label_cname2ename供后续使用
    label_cname2ename = get_st_table(label_id2cid, label_id2value, label_cid2cname)

    # 处理connectionCategories.json文件
    connect_cid2cname = {}
    with open(connect_path_c, "r", encoding="utf-8") as f:
        json_list = json.load(f)
    for item_dict in json_list:
        key = item_dict['id']
        connect_cid2cname[str(key)] = item_dict['text']

    # 处理connections.json文件
    bt_connect = []      # 生成'本体'表格关系部分所需列表
    connect_cname2ename = {}
    connect_id2all = {}  # 生成'关系'表格所需字典
    with open(connect_path, "r", encoding="utf-8") as f:
        json_list = json.load(f)
    for item_dict in json_list:
        all_list = []
        # 确定from_id
        from_id = str(item_dict['srcId']) + '-' + str(item_dict['fromId'])  # modify by 20220627
        all_list.append(from_id)
        # 确定to_id
        to_id = str(item_dict['srcId']) + '-' + str(item_dict['toId'])  # modify by 20220627
        all_list.append(to_id)
        # 确定from_label的ename
        from_label = label_cname2ename[label_cid2cname[str(label_id2cid[from_id])]]
        # from_label = label_cname2ename[label_cid2cname[str(label_id2cid[from_id])]]
        all_list.append(from_label)
        # 确定to_label的ename
        to_label = label_cname2ename[label_cid2cname[str(label_id2cid[to_id])]]
        # to_label = label_cname2ename[label_cid2cname[str(label_id2cid[to_id])]]
        all_list.append(to_label)
        # 确定connect_label, 并返回一个列表bt_connect
        connect_cid = str(item_dict['categoryId'])  # modify by 20220627
        connect_label = from_label + '_' + to_label + '_' + str(connect_cid)
        all_list.append(connect_label)
        # 以下子部分为了得到列表bt_connect
        connect_cname = connect_cid2cname[str(connect_cid)]
        if connect_cname not in connect_cname2ename.keys():
            connect_cname2ename[connect_cname] = []
            connect_cname2ename[connect_cname].append(connect_label)
            bt_connect.append([connect_label, connect_cname, from_label, to_label])
        elif connect_label not in connect_cname2ename[connect_cname]:
            connect_cname2ename[connect_cname].append(connect_label)
            bt_connect.append([connect_label, connect_cname, from_label, to_label])
        # 确定最终的value, 生成字典connect_id2all
        key = str(item_dict['srcId']) + '-' + str(item_dict['id'])  # modify by 20220627
        connect_id2all[key] = all_list
        del all_list

    # 清理中间数据
    del connect_cname2ename

    # 生成‘关系’表格
    get_gx_table(connect_id2all)

    # 生成‘本体’表格
    get_bt_table(label_cname2ename, bt_connect)

    print('完成!!!')


def ud2table(folder, output):
    if os.path.exists(folder) and os.path.isdir(folder):
        label_path = os.path.join(folder, "labels.json")
        connect_path = os.path.join(folder, "connections.json")
        label_path_c = os.path.join(folder, "labelCategories.json")
        connect_path_c = os.path.join(folder, "connectionCategories.json")

        if os.path.exists(label_path) and os.path.exists(connect_path) and os.path.exists(label_path_c) and os.path.exists(connect_path_c):
            # 文件均存在则开始转换
            handle_json_file(label_path, connect_path, label_path_c, connect_path_c, output)
        else:
            print('一些json文件不存在或提供的是目录!!!')
            print('必须提供一下4份json文件: labels.json, connections.json, labelCategories.json, connectionCategories.json')
    else:
        result_json = {}
        result_json["error_code"] = 1001
        result_json["error_msg"] = ("不是文件夹或不存在 %s" % folder)
        print(json.dumps(result_json, ensure_ascii=False))
        sys.exit(1)


if __name__ == '__main__':
    print("start")
