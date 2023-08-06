"""
Required: pip install openpyxl defusedxml
"""
from __future__ import annotations
import logging, sys
from typing import Any
from datetime import datetime, timezone
from pathlib import Path

if sys.version_info[0:2] < (3, 8):
    from typing_extensions import Literal
else:
    from typing import Literal

from . import filemgr
from .base import BaseOut

try:
    from openpyxl import load_workbook, Workbook, DEFUSEDXML
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo
    from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
    from openpyxl.styles.fills import PatternFill
    from openpyxl.utils import range_boundaries, get_column_letter

    _import_error = None

    logger = logging.getLogger(__name__)

    _cache = {
        'defusedxml_alert': False,
    }


    class ExcelOut(BaseOut):
        @classmethod
        def is_available(cls):
            return _import_error is None
        

        def __init__(self, timezone: Literal['local']|timezone = None, **kwargs):
            super().__init__(**kwargs)

            self._timezone = timezone

            if not self.is_available():
                raise ValueError(f"cannot use {self.__class__.__name__}: {_import_error}")

            if not DEFUSEDXML and not _cache['defusedxml_alert']:
                logger.warning("By default openpyxl does not guard against quadratic blowup or billion laughs xml attacks. To guard against these attacks install defusedxml.")
                _cache['defusedxml_alert'] = True

            if not self._table_name:
                self._table_name = self._DEFAULT_TABLE_NAME


        def _format_value(self, value: Any) -> str:
            if value is None:
                return None
            elif isinstance(value, datetime) and value.tzinfo:
                # Excel does not support timezones in datetimes
                if self._timezone:
                    value = value.astimezone(None if self._timezone == 'local' else self._timezone)
                return value.replace(tzinfo=None)
            else:
                return super()._format_value(value)


        def _open_path(self):                
            if not self._path:
                raise ValueError(f'cannot open path for {self._name}')
                
            if not hasattr(self, '_workbook'):
                self._read_existing_file_if_any()

            # Open path for binary writting
            return filemgr.open_file(self._path, mode='wb', newline=self._newline, encoding=self._encoding, mkdir=True)


        _workbook: Workbook
        _worksheet: Worksheet
        _previous_table: Table|None
        
        def _read_existing_file_if_any(self):
            # Open workbook and search for existing table
            if filemgr.exists(self._path):
                with filemgr.open_file(self._path, 'rb') as fd:
                    self._workbook: Workbook = load_workbook(fd)

                self._previous_table = None
                for name in self._workbook.sheetnames:
                    self._worksheet: Worksheet = self._workbook[name]
                    if self._table_name in self._worksheet.tables:
                        self._previous_table = self._worksheet.tables[self._table_name]
                        break

                if not self._previous_table:
                    # table not found: we create a new worksheet
                    self._worksheet: Worksheet = self._workbook.create_sheet(title=self._table_name)

            else:
                self._workbook: Workbook = Workbook()
                self._worksheet: Worksheet = self._workbook.active
                self._worksheet.title = self._table_name
                self._previous_table = None


        _headers_mapping: dict[int,int]|None
        _unmanaged_col_indices: list[int]|None
        _column_formats: dict[int,dict[str,Any]]|None
        _previous_last_row_index: int|None
        _first_col_index: int
        _first_row_index: int
        _last_col_index: int
        _next_row_index: int
        _warned_additional_columns: bool


        def _output_headers(self):
            self._headers_mapping = None
            self._unmanaged_col_indices = None
            self._column_formats = None

            if not hasattr(self, '_workbook'):
                self._read_existing_file_if_any()
            
            if self._previous_table:
                # Get existing table boundaries
                self._first_col_index, self._first_row_index, self._last_col_index, self._previous_last_row_index = range_boundaries(self._previous_table.ref)

                # Determines previous headers   
                previous_headers: list[str] = []
                column: TableColumn
                for i, column in enumerate(self._previous_table.tableColumns):
                    previous_headers.append(column.name)
                    self._prepare_column_format(col_index=self._first_col_index + i, column=column)

                # Determines headers mapping
                new_headers = []
                for pos, header in enumerate(self._headers):
                    try:
                        previous_pos = previous_headers.index(header)
                    except ValueError:
                        new_headers.append(header)
                        previous_pos = len(previous_headers) + len(new_headers) - 1

                    if previous_pos != pos:
                        if self._headers_mapping is None:
                            self._headers_mapping = {}
                        self._headers_mapping[pos] = previous_pos

                # Determines header columns that are not handled by us
                for pos, header in enumerate(previous_headers):
                    if not header in self._headers:
                        if self._unmanaged_col_indices is None:
                            self._unmanaged_col_indices = []
                        self._unmanaged_col_indices.append(self._first_col_index + pos)
                
                # Add new headers
                if new_headers:
                    logger.info(f"add header for table {self._table_name}: {', '.join(new_headers)} - was missing in existing table")

                    for header in new_headers:
                        self._last_col_index += 1
                        self._set_cell(self._first_row_index, self._last_col_index, header, is_header=True)

                        # Mark data as '?' (unknown) for the column if we append
                        if self._append:
                            r = self._first_row_index + 1
                            while r <= self._previous_last_row_index:
                                self._set_cell(r, self._last_col_index, '?')
                                r += 1

                # Determine next row
                if self._append:
                    self._next_row_index = self._previous_last_row_index + 1
                else:
                    self._next_row_index = self._first_row_index + 1
            
            else:
                # Table does not exist: write headers
                self._previous_last_row_index = None
                self._first_col_index = 1
                self._first_row_index = 1
                self._last_col_index = 0

                for header in self._headers:
                    self._last_col_index += 1
                    self._set_cell(self._first_col_index, self._last_col_index, header, is_header=True)

                self._next_row_index = self._first_row_index + 1

            self._warned_additional_columns = False


        def _on_append(self, row: list):
            if not self._headers:
                raise ValueError(f"headers not appended yet")

            # Write managed data
            for i in range(0, len(row)):
                if i >= len(self._headers):
                    if not self._warned_additional_columns:
                        logger.warning(f"ignore row values at index >= {len(self._headers)} (first occurence on row {len(self.rows)})")
                        self._warned_additional_columns = True
                else:
                    col_index = self._first_col_index + (self._headers_mapping.get(i, i) if self._headers_mapping else i)
                    self._set_cell(self._next_row_index, col_index, row[i])

            # Erase cell for unmanaged data
            if self._unmanaged_col_indices:
                for col_index in self._unmanaged_col_indices:
                    self._erase_cell(self._next_row_index, col_index)

            self._next_row_index += 1
        
        
        def _before_close(self):
            if not self._path:
                raise ValueError(f"can only use ExcelOut with path target")

            if not self._headers:
                raise ValueError(f"cannot use ExcelOut without headers")

            # Erase data outside of table
            if self._previous_last_row_index is not None:
                r = self._next_row_index
                while r <= self._previous_last_row_index:
                    c = self._first_col_index
                    while c <= self._last_col_index:
                        self._erase_cell(r, c, outside_table = True)
                        c += 1
                    r += 1

            # Create table
            table_ref = f"{get_column_letter(self._first_col_index)}{self._first_row_index}:{get_column_letter(self._last_col_index)}{self._next_row_index-1}"
            if self._previous_table:
                if self._previous_table.ref != table_ref:
                    self._recreate_table(table_ref)
            else:
                table = Table(name=self._table_name, displayName=self._table_name, ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                self._worksheet.add_table(table)

            # Save the file
            self._workbook.save(self.file)


        def _recreate_table(self, new_ref):
            newcolumns = []

            for i in range(0, self._last_col_index - self._first_col_index + 1):
                name = self._worksheet.cell(self._first_row_index, self._first_col_index + i).value
                newcolumn = TableColumn(id=i+1, name=name)
                newcolumns.append(newcolumn)

                if i < len(self._previous_table.tableColumns):
                    prevcolumn: TableColumn = self._previous_table.tableColumns[i]
                    newcolumn.dataCellStyle = prevcolumn.dataCellStyle
                    newcolumn.dataDxfId = prevcolumn.dataDxfId # refers to workbook._differential_styles
                    newcolumn.calculatedColumnFormula = prevcolumn.calculatedColumnFormula

            newtable = Table(name=self._table_name, displayName=self._table_name, ref=new_ref, tableColumns=newcolumns, autoFilter=self._previous_table.autoFilter, sortState=self._previous_table.sortState)
            newtable.tableStyleInfo = self._previous_table.tableStyleInfo
            
            del self._worksheet.tables[self._table_name]
            self._worksheet.add_table(newtable)


        # -------------------------------------------------------------------------
        # Helpers
        # -------------------------------------------------------------------------
        _DEFAULT_TABLE_NAME = 'Flexout'

        def _set_cell(self, row_index: int, col_index: int, value, is_header = False) -> Cell:
            cell: Cell = self._worksheet.cell(row_index, col_index)

            if not is_header:
                self._apply_column_format(cell)

            try:
                cell.value = value
            except ValueError as err:
                if str(err).startswith('Cannot convert'):
                    cell.value = str(value)
                else:
                    raise

            return cell


        def _erase_cell(self, row_index: int, col_index: int, outside_table = False) -> Cell:
            cell: Cell = self._worksheet.cell(row_index, col_index)
            cell.style = 'Normal'
            
            cell.value = None
            if not outside_table:
                self._apply_column_format(cell)
            
            return cell


        def _apply_column_format(self, cell: Cell):
            if not self._column_formats:
                return

            fmt = self._column_formats.get(cell.col_idx, None)
            if fmt is None:
                return

            if 'formula' in fmt:
                formula = fmt['formula']
                if isinstance(formula, ArrayFormula):
                    pass # TODO: not supported yet
                else:
                    cell.value = formula

            if 'style' in fmt:
                cell.style = fmt['style']

            for fmt_key, fmt_value in fmt.items():
                if fmt_key in ['formula', 'style']:
                    continue
                setattr(cell, fmt_key, fmt_value)


        def _prepare_column_format(self, col_index: int, column: TableColumn) -> dict[str,Any]|None:
            if not self._column_formats:
                self._column_formats = {}

            fmt: dict[str,Any] = None

            # Read dataCellStyle
            if column.dataCellStyle:
                if fmt is None:
                    fmt = {}

                fmt['style'] = column.dataCellStyle
            
            # Read dxf
            if column.dataDxfId is not None:
                if fmt is None:
                    fmt = {}

                dxf: DifferentialStyle = self._workbook._differential_styles[column.dataDxfId]

                if dxf.numFmt:
                    fmt['number_format'] = dxf.numFmt.formatCode
                else:
                    if not 'style' in fmt:
                        fmt['number_format'] = self._DEFAULT_NUMBER_FORMAT

                fmt['alignment'] = dxf.alignment if dxf.alignment else self._DEFAULT_ALIGNMENT
                fmt['border'] = dxf.border if dxf.border else self._DEFAULT_BORDER
                fmt['font'] = dxf.font if dxf.font else self._DEFAULT_FONT
                fmt['protection'] = dxf.protection if dxf.protection else self._DEFAULT_PROTECTION
                fmt['fill'] = PatternFill(fill_type=dxf.fill.fill_type, bgColor=dxf.fill.fgColor, fgColor=dxf.fill.bgColor) if dxf.fill else self._DEFAULT_FILL # NOTE: fgcolor and bgcolor are inversed in DifferentialStyle

            # Read formula
            if column.calculatedColumnFormula:
                if fmt is None:
                    fmt = {}

                formula = column.calculatedColumnFormula
                if formula.array:
                    fmt['formula'] = ArrayFormula(formula.attr_text)
                else:
                    fmt['formula'] = '=' + formula.attr_text
            
            # Register format
            if fmt is not None:
                self._column_formats[col_index] = fmt
        

        _DEFAULT_NUMBER_FORMAT = 'General'

        _DEFAULT_FILL = PatternFill(fill_type=None)

        _DEFAULT_ALIGNMENT = None # openpyxl.styles.alignment.Alignment
        _DEFAULT_BORDER = None # openpyxl.styles.alignment.Border
        _DEFAULT_FONT = None # openpyxl.styles.fonts.Font
        _DEFAULT_PROTECTION = None # openpyxl.styles.protection.Protection


# -----------------------------------------------------------------------------
except ImportError as err:
    _import_error = str(err)

    class ExcelOut(BaseOut):
        @classmethod
        def is_available(cls):
            return False
        
        def __init__(self, **kwargs):
            super().__init__(**kwargs)

            if not self.is_available():
                raise ValueError(f"cannot use {self.__class__.__name__}: {_import_error}")
