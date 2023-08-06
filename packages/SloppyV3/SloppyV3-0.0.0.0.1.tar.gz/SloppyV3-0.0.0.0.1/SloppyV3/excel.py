#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
# File: excel.py
# Project: SloppyV3
# Created Date: 2023-02-13, 10:13:27
# Author: Chungman Kim(h2noda@unipark.kr)
# Last Modified: Mon Mar 06 2023
# Modified By: Chungman Kim
# Copyright 2023. Chungman Kim
# HISTORY:
# Date      	By	Comments
# ----------	---	----------------------------------------------------------
'''

from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, numbers, borders
from openpyxl import Workbook


class Xlsx:
    def __init__(self, filename, sheetname="sheet", mode="n"):
        """Mode = n : New Excel File
                  r : Read Excel File

        Args:
            filename (String): Excel File Name
            sheetname (String): Excel Sheet Name
            mode (str, optional): Excel file Mode, Defaults to "n".
        """
        self.filename = filename
        self.sheetname = sheetname
        if (mode == "n"):
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.title = sheetname
        elif (mode == "r"):
            self.wb = openpyxl.load_workbook(filename)
            self.sheet = self.wb[sheetname]

    def SideStyle(border_style=borders.NONE, color="000000"):
        side = Side(border_style=border_style, color=color)
        return side

    @dataclass
    class CellStyle:
        # row: int
        # column: int
        # cellvalue: str = None
        font: Font = Font(name="D2Coding", size="10", color="000000",
                          underline=None, bold=False, italic=False, strikethrough=False)
        alignment: Alignment = Alignment(
            horizontal="center", vertical="center")
        side: Side(border_style=borders.NONE, color="000000")
        border: Border = Border(left=side, right=side, top=side, bottom=side)
        fill: PatternFill = PatternFill(fill_type="solid",
                                        start_color=None, end_color=None)

        def set_font(name="D2Coding", size="10", color="000000",
                          underline=None, bold=False, italic=False, strikethrough=False) -> Font:
            font = Font(name=name, size=size, color=color, underline=underline,
                        bold=bold, italic=italic, strikethrough=strikethrough)
            return font

        def set_alignment(horizontal="center", vertical="center") -> Alignment:
            alignment = Alignment(horizontal=horizontal, vertical=vertical)
            return alignment

        def set_side(border_style=borders.BORDER_NONE, color="000000") -> Side:
            side = Side(border_style=border_style, color=color)
            return side
        # def set_border()

    def create_sheet(self, sheetname, loc=0):
        """Create Excel Sheet

        Args:
            sheetname (String): Sheetname
            loc (int, optional): Sheet no. Defaults to 0(first).
        """
        self.sheet = self.wb.create_sheet(sheetname, loc)

    def select_sheet(self, sheetname):
        """Select Excel Sheet

        Args:
            sheetname (String): Sheetname
        """
        self.sheet = self.wb[sheetname]

    def rename_sheetname(self, before_name, after_name):
        """Rename Excel Sheet Name

        Args:
            before_name (String): Before Sheet Name
            after_name (String): After Sheet Name
        """
        self.wb[before_name].title = after_name

    def save(self):
        """Save Excel file
        """
        self.wb.save(self.filename)
