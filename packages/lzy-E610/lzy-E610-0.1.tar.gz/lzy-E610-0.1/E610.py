import re

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter


class E610:
    def __init__(self, **kwargs):
        self.filepath = kwargs.get('filepath')
        if self.filepath:
            self.workbook = openpyxl.load_workbook(filename=self.filepath, data_only=kwargs.get('data_only'))
        else:
            self.workbook = openpyxl.Workbook()

    def read_sheet(self, close=True, **kwargs):
        if kwargs.get('sheet_name'):
            self.active_sheet = self.workbook[kwargs.get('sheet_name')]
        else:
            self.active_sheet = self.workbook.active
        coordinate = kwargs.get('coordinate')
        start_column = int(column_index_from_string(re.findall(r'([a-z]+)\d+:([a-z]+)\d+', coordinate, re.I)[0][0]))
        end_column = int(column_index_from_string(re.findall(r'([a-z]+)\d+:([a-z]+)\d+', coordinate, re.I)[0][1]))
        start_row = int(re.findall(r'[a-z]+(\d+):[a-z]+(\d+)', coordinate, re.I)[0][0])
        end_row = int(re.findall(r'[a-z]+(\d+):[a-z]+(\d+)', coordinate, re.I)[0][1])
        value_list = []
        if kwargs.get('iter').lower() == 'row':
            for row_cell in self.active_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column,
                                                        max_col=end_column):
                if kwargs.get('value') == True:
                    value_list.append([cell.value for cell in row_cell])
                else:
                    value_list.append([cell for cell in row_cell])
        elif kwargs.get('iter').lower() == 'column':
            for column_cell in self.active_sheet.iter_cols(min_row=start_row, max_row=end_row, min_col=start_column,
                                                           max_col=end_column):
                if kwargs.get('value') == True:
                    value_list.append([cell.value for cell in column_cell])
                else:
                    value_list.append([cell for cell in column_cell])
        if close:
            self.workbook.close()
        return value_list

    def write_sheet(self, save_close=True, **kwargs):
        if kwargs.get('sheet_name'):
            self.active_sheet = self.workbook[kwargs.get('sheet_name')]
        else:
            self.active_sheet = self.workbook.active
        coordinate = kwargs.get('coordinate')
        start_column = int(column_index_from_string(re.findall(r'([a-z]+)', coordinate, re.I)[0]))
        start_row = int(re.findall(r'(\d+)', coordinate, re.I)[0])
        for row in range(len(kwargs.get('list'))):
            for column in range(len(kwargs.get('list')[row])):
                self.active_sheet.cell(row + start_row, column + start_column).value = kwargs.get('list')[row][column]
        if save_close:
            self.workbook.save(self.filepath)
            self.workbook.close()

    def max_row_or_column(self, **kwargs):
        if kwargs.get('sheet_name'):
            self.active_sheet = self.workbook[kwargs.get('sheet_name')]
        else:
            self.active_sheet = self.workbook.active
        if kwargs.get('value'):
            if kwargs.get('value').isdigit():
                return max((bb.column for bb in self.active_sheet[kwargs.get('value')] if bb.value))
            elif kwargs.get('value').isalpha():
                return max((bb.row for bb in self.active_sheet[kwargs.get('value')] if bb.value))
        else:
            return get_column_letter(self.active_sheet.max_column), self.active_sheet.max_row
